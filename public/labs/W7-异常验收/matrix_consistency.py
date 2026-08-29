import argparse
import csv
import re
import sys
import xml.etree.ElementTree as ET
from pathlib import Path
from zipfile import ZipFile


BASE_DIR = Path(__file__).resolve().parent
CSV_PATH = BASE_DIR / "acceptance-matrix.csv"
XLSX_PATH = BASE_DIR / "acceptance-matrix.xlsx"

CSV_HEADERS = [
    "case_id",
    "layer",
    "precondition",
    "action",
    "expected_http",
    "expected_business",
    "evidence",
    "actual_result",
    "status",
    "review_notes",
]
XLSX_HEADERS = [
    "用例ID",
    "层级",
    "前置条件",
    "操作",
    "预期HTTP",
    "预期业务结果",
    "证据",
    "实际结果",
    "状态",
    "复盘备注",
]
XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def load_csv_rows() -> list[list[str]]:
    with CSV_PATH.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != CSV_HEADERS:
            raise ValueError(
                f"CSV 列不符合权威结构：\n当前={reader.fieldnames}\n期望={CSV_HEADERS}"
            )
        rows = [[row[header] for header in CSV_HEADERS] for row in reader]

    case_ids = [row[0] for row in rows]
    if len(rows) != 20:
        raise ValueError(f"权威 CSV 应为 20 条，当前为 {len(rows)} 条")
    if len(set(case_ids)) != len(case_ids):
        raise ValueError("权威 CSV 存在重复用例 ID")
    return rows


def column_index(reference: str) -> int:
    letters = re.match(r"[A-Z]+", reference)
    if letters is None:
        raise ValueError(f"无法识别 XLSX 单元格引用 {reference!r}")
    value = 0
    for letter in letters.group(0):
        value = value * 26 + ord(letter) - ord("A") + 1
    return value - 1


def shared_strings(archive: ZipFile) -> list[str]:
    path = "xl/sharedStrings.xml"
    if path not in archive.namelist():
        return []
    root = ET.fromstring(archive.read(path))
    return ["".join(node.itertext()) for node in root.findall(f"{{{XML_NS}}}si")]


def cell_text(cell: ET.Element, strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        inline = cell.find(f"{{{XML_NS}}}is")
        return "" if inline is None else "".join(inline.itertext())
    value = cell.find(f"{{{XML_NS}}}v")
    if value is None or value.text is None:
        return ""
    if cell_type == "s":
        return strings[int(value.text)]
    return value.text


def load_xlsx_table() -> tuple[list[str], list[list[str]]]:
    with ZipFile(XLSX_PATH) as archive:
        strings = shared_strings(archive)
        root = ET.fromstring(archive.read("xl/worksheets/sheet1.xml"))

    parsed_rows: list[list[str]] = []
    for row in root.findall(f".//{{{XML_NS}}}row"):
        values = [""] * len(XLSX_HEADERS)
        for cell in row.findall(f"{{{XML_NS}}}c"):
            index = column_index(cell.attrib["r"])
            if index < len(values):
                values[index] = cell_text(cell, strings)
        parsed_rows.append(values)

    try:
        header_index = parsed_rows.index(XLSX_HEADERS)
    except ValueError as exc:
        raise ValueError("XLSX 中找不到完整的 10 列验收表头") from exc
    data_rows = [row for row in parsed_rows[header_index + 1 :] if row[0]]
    return parsed_rows[header_index], data_rows


def write_xlsx(csv_rows: list[list[str]]) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "--write 需要 openpyxl；可先安装，或只运行无依赖的 --check"
        ) from exc

    workbook = load_workbook(XLSX_PATH)
    sheet = workbook[workbook.sheetnames[0]]
    header_row = None
    for row_number in range(1, sheet.max_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(1, 11)]
        if values == XLSX_HEADERS:
            header_row = row_number
            break
    if header_row is None:
        raise ValueError("XLSX 中找不到完整的 10 列验收表头")

    first_data_row = header_row + 1
    for row_offset, values in enumerate(csv_rows):
        for column, value in enumerate(values, start=1):
            cell_value: str | int | None = value
            if column == 5 and value.isdigit():
                cell_value = int(value)
            elif value == "":
                cell_value = None
            sheet.cell(first_data_row + row_offset, column).value = cell_value

    for row_number in range(first_data_row + len(csv_rows), sheet.max_row + 1):
        for column in range(1, 11):
            sheet.cell(row_number, column).value = None
    workbook.save(XLSX_PATH)


def check(csv_rows: list[list[str]]) -> None:
    headers, xlsx_rows = load_xlsx_table()
    if headers != XLSX_HEADERS:
        raise ValueError("XLSX 表头与预期不一致")
    if len(xlsx_rows) != len(csv_rows):
        raise ValueError(f"CSV/XLSX 行数不一致：{len(csv_rows)} != {len(xlsx_rows)}")

    differences: list[str] = []
    for row_number, (csv_row, xlsx_row) in enumerate(zip(csv_rows, xlsx_rows), start=1):
        for column, (csv_value, xlsx_value) in enumerate(zip(csv_row, xlsx_row), start=1):
            if csv_value != xlsx_value:
                differences.append(
                    f"第 {row_number} 条 {csv_row[0]} / {XLSX_HEADERS[column - 1]}："
                    f"CSV={csv_value!r}，XLSX={xlsx_value!r}"
                )
    if differences:
        preview = "\n".join(f"- {item}" for item in differences[:20])
        suffix = "" if len(differences) <= 20 else f"\n... 共 {len(differences)} 处差异"
        raise ValueError(f"CSV/XLSX 内容漂移：\n{preview}{suffix}")


def main() -> int:
    parser = argparse.ArgumentParser(description="同步或检查 W7 验收矩阵")
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument("--write", action="store_true", help="用权威 CSV 覆盖 XLSX 的 10 列内容")
    mode.add_argument("--check", action="store_true", help="检查 CSV 与 XLSX 是否完全一致（默认）")
    args = parser.parse_args()

    try:
        csv_rows = load_csv_rows()
        if args.write:
            write_xlsx(csv_rows)
            print(f"已从权威 CSV 同步 {len(csv_rows)} 条用例到 XLSX")
        check(csv_rows)
    except (OSError, ValueError, RuntimeError) as exc:
        print(f"FAIL {exc}", file=sys.stderr)
        return 1

    print(f"PASS CSV/XLSX 完全一致：{len(csv_rows)} 条 × {len(CSV_HEADERS)} 列")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
